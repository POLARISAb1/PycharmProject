# coding=utf-8

# 获取工作簿
from openpyxl import load_workbook

book = load_workbook('test.xlsx')
print(book.sheetnames)

# 读取工作表
sheet = book.active

# 读取单个单元格
cell_a1 = sheet['A1']
cell_a3 = sheet.cell(column=1, row=3)

# 读取多个连续单元格
# cells = sheet.iter_rows(max_row=3, max_col=2, values_only=True)
# print(type(cells), cells)
# print(isinstance(cells, Iterable))
# for row in cells:
#     print(type(row), row)
#     # 继续遍历
#     for cc in row:
#         print(type(cc), cc)

#
# r = sheet.rows
# print(r)
#
# # for cell in r:
# #     print(cell)
# # 创建并返回一个新对象
# cl = tuple(r)
# print(cl)

# 获取单元格的值

# a_1 = sheet['A1']
# print(a_1.value)


# all_s = sheet.values
# print(all_s)
# for row in all_s:
#     # print(type(row), row)
#     for cell in row :
#         print(cell)
